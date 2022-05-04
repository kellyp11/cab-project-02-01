<<<<<<< HEAD
<<<<<<<< HEAD:doc/5a/Scripts/convert-CSV-collab-V2.py
import pandas as pd
import numpy as np
import openpyxl

file_loc = "ESPA_Data_CAB.xlsx" #File name from Mr. Paul Romano
newSheet = pd.ExcelWriter('Updated_ESPA_Data_CAB.xlsx', engine='xlsxwriter') #File name that is to be created

ME_CONSTANT = "Meter Entries"
M_CONSTANT = "Meters"

specialCols_DI = "F, G, H" #Columns for the DATE_INTERVAL table
sheetName_DI = 'Meter Entries' #Sheet for the DATE_INTERVAL table

specialCols_B = "B, A, L, M" #Columns for the BUILDING table
sheetName_B = 'Properties' # Sheet for the BUILDING table

specialCols_BT = "A, K"  #Columns for the BUILDING_TYPE table
sheetName_BT = 'Properties' #Sheet for the BUILDING_TYPE table

specialCols_ES = "C, D, E" #Columns for the ENERGY_SOURCE table
sheetName_ES = M_CONSTANT #Sheet for the ENERGY_SOURCE table

specialCols_ESC = "C, L, J" #Columns for the ENERGY_SOURCE_COST table
sheetName_ESC = ME_CONSTANT #Sheet for the ENERGY_SOURCE_COST table

specialCols_FO = "C, F, E" #Columns for the FUEL_OIL table
sheetName_FO = M_CONSTANT #Sheet for the FUEL_OIL table

specialCols_NG = "C, F, E" #Columns for the NATURAL_GAS table
sheetName_NG = M_CONSTANT #Sheet for the NATURAL_GAS table

specialCols_EG = "C, F, E" #Columns for the ELECTRIC_GRID table
sheetName_EG = M_CONSTANT #Sheet for the ELECTRIC_GRID table

specialCols_OS = "C, F, E" #Columns for the OTHER_SOURCE table
sheetName_OS = M_CONSTANT #Sheet for the OTHER_SOURCE table

specialCols_MT = "F, C" #Columns for the MAPS_TO table
sheetName_MT = 'Meter Entries' #Sheet for the MAPS_TO table

specialCols_PB = "C, B" #Columns for the POWERED_BY table
sheetName_PB = 'Meters' #Sheet for the POWERED_BY table


#This is the code to handle the manipulation of Date Interval

#Start Date:
idf = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_DI, usecols = 'G, H')

idf[['start','end']] = idf[['Start Date','End Date']].apply(pd.to_datetime)
idf['start'] = idf['start'].dt.strftime('%Y-%m-%d')

s = pd.concat([pd.Series(r.Index, pd.date_range(r.start, r.end, freq = '15T').strftime('%Y-%m-%d')) 
                         for r in idf.itertuples()])       

dates = idf.loc[s].drop(['end'], axis=1).assign(start=s.index).reset_index(drop=True)

#dates = idf['start'].dt.strftime('%Y-%m-%d')
#dates = (pd.DataFrame(columns=['StartDate'],
       #index=pd.date_range(str(idf.iloc[0,0]) + 'T00:00:00', str(idf.iloc[0,1]) + 'T23:59:00', 
              #freq='15T'))
              #.between_time('00:00','23:59')
              #.index.strftime('%Y-%m-%d'))
       
#print(dates)

#Intervals:

intervals = (pd.DataFrame(columns=['StartTimestamp'],
                  index=pd.date_range(str(idf.iloc[0,0]) + 'T00:00:00', str(idf.iloc[0,1]) + 'T23:59:00',
                                      freq='15T'))
       .between_time('00:00','23:59')
       .index.strftime('%H:%M:%S'))
#print(intervals)

#MeterConsumptionID:
meterConsumpID = pd.DataFrame(columns = ['Meter Consumption ID'])

# INCOMPLETE
#for x in range(2880):
#    idf.loc[meterConsumpID,'Meter Consumption ID'] = (meterConsumpID + '-' + x).zfill(len(4))

dates_df = pd.DataFrame (dates, columns = ['start'])
intervals_df = pd.DataFrame (intervals, columns = ['StartTimestamp'])
id_key = pd.DataFrame(meterConsumpID, columns = ['Meter Consumption ID'])

#print(timestamps)
df_timestamps = pd.concat([dates_df, intervals_df, id_key], axis=1)

df_timestamps.to_csv('DI.csv', index = None, header = True)

#df_timestamps.to_excel(newSheet, sheet_name= 'DATE_INTERVAL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)
#idf.to_excel(newSheet, sheet_name= 'DATE_INTERVAL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=1, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

# --------------------------

#Modifying the sheet for POWERED_BY table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_PB, usecols = specialCols_PB)
#print(df)
df.to_excel(newSheet, sheet_name= 'POWERED_BY', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

# --------------------------

#Modifying the sheet for BUILDING table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_B, usecols = specialCols_B)
#print(df)
df.to_excel(newSheet, sheet_name= 'BUILDING', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for BUILDING_TYPE table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_BT, usecols = specialCols_BT)
#print(df)
df.to_excel(newSheet, sheet_name= 'BUILDING_TYPE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ENERGY_SOURCE
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_ES, usecols = specialCols_ES)
#print(df)
df.to_excel(newSheet, sheet_name= 'ENERGY_SOURCE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ENERGY_SOURCE_COST table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_ESC, usecols = specialCols_ESC)
#print(df)
df.to_excel(newSheet, sheet_name= 'ENERGY_SOURCE_COST', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for FUEL_OIL table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_FO, usecols = specialCols_FO)
#print(df)
df.to_excel(newSheet, sheet_name= 'FUEL_OIL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for NATURAL_GAS table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_NG, usecols = specialCols_NG)
#print(df)
df.to_excel(newSheet, sheet_name= 'NATURAL_GAS', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ELECTRIC_GRID table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_EG, usecols = specialCols_EG)
#print(df)
df.to_excel(newSheet, sheet_name= 'ELECTRIC_GRID', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for OTHER_SOURCE table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_OS, usecols = specialCols_OS)
#print(df)
df.to_excel(newSheet, sheet_name= 'OTHER_SOURCE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for MAPS_TO table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_MT, usecols = specialCols_MT)
#print(df)
df.to_excel(newSheet, sheet_name= 'MAPS_TO', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

newSheet.save()
========
=======
<<<<<<< HEAD
>>>>>>> f7fe8a7 (Updating Branch)
import pandas as pd
import numpy as np
import openpyxl

file_loc = "ESPA_Data_CAB.xlsx" #File name from Mr. Paul Romano
newSheet = pd.ExcelWriter('Updated_ESPA_Data_CAB.xlsx', engine='xlsxwriter') #File name that is to be created

ME_CONSTANT = "Meter Entries"
M_CONSTANT = "Meters"

specialCols_DI = "F, G, H" #Columns for the DATE_INTERVAL table
sheetName_DI = 'Meter Entries' #Sheet for the DATE_INTERVAL table

specialCols_B = "B, A, L, M" #Columns for the BUILDING table
sheetName_B = 'Properties' # Sheet for the BUILDING table

specialCols_BT = "A, K"  #Columns for the BUILDING_TYPE table
sheetName_BT = 'Properties' #Sheet for the BUILDING_TYPE table

specialCols_ES = "C, D, E" #Columns for the ENERGY_SOURCE table
sheetName_ES = M_CONSTANT #Sheet for the ENERGY_SOURCE table

specialCols_ESC = "C, L, J" #Columns for the ENERGY_SOURCE_COST table
sheetName_ESC = ME_CONSTANT #Sheet for the ENERGY_SOURCE_COST table

specialCols_FO = "C, F, E" #Columns for the FUEL_OIL table
sheetName_FO = M_CONSTANT #Sheet for the FUEL_OIL table

specialCols_NG = "C, F, E" #Columns for the NATURAL_GAS table
sheetName_NG = M_CONSTANT #Sheet for the NATURAL_GAS table

specialCols_EG = "C, F, E" #Columns for the ELECTRIC_GRID table
sheetName_EG = M_CONSTANT #Sheet for the ELECTRIC_GRID table

specialCols_OS = "C, F, E" #Columns for the OTHER_SOURCE table
sheetName_OS = M_CONSTANT #Sheet for the OTHER_SOURCE table

specialCols_MT = "F, C" #Columns for the MAPS_TO table
<<<<<<< HEAD
sheetName_MT = ME_CONSTANT #Sheet for the MAPS_TO table

specialCols_PB = "C, B" #Columns for the POWERED_BY table
sheetName_PB = M_CONSTANT #Sheet for the POWERED_BY table

#This is the code to handle the manipulation of Date Interval
#This does not work so we're coming back to this later.
#The format of the data works outside of the relations
#DATE_INTERVAL and MAPS_TO works perfectly fine.

idf = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_DI, usecols = 'G, H')

dates = (pd.DataFrame(columns=['StartDate'],
                  index=pd.date_range(str(idf.iloc[0,0]) + 'T00:00:00', str(idf.iloc[0,1]) + 'T23:59:00',
                                      freq='15T'))
       .between_time('00:00','23:59')
       .index.strftime('%Y-%m-%d'))
       
#print(dates)

=======
sheetName_MT = 'Meter Entries' #Sheet for the MAPS_TO table

specialCols_PB = "C, B" #Columns for the POWERED_BY table
sheetName_PB = 'Meters' #Sheet for the POWERED_BY table


#This is the code to handle the manipulation of Date Interval

#Start Date:
idf = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_DI, usecols = 'G, H')

idf[['start','end']] = idf[['Start Date','End Date']].apply(pd.to_datetime)
idf['start'] = idf['start'].dt.strftime('%Y-%m-%d')

s = pd.concat([pd.Series(r.Index, pd.date_range(r.start, r.end, freq = '15T').strftime('%Y-%m-%d')) 
                         for r in idf.itertuples()])       

dates = idf.loc[s].drop(['end'], axis=1).assign(start=s.index).reset_index(drop=True)

#dates = idf['start'].dt.strftime('%Y-%m-%d')
#dates = (pd.DataFrame(columns=['StartDate'],
       #index=pd.date_range(str(idf.iloc[0,0]) + 'T00:00:00', str(idf.iloc[0,1]) + 'T23:59:00', 
              #freq='15T'))
              #.between_time('00:00','23:59')
              #.index.strftime('%Y-%m-%d'))
       
#print(dates)

#Intervals:

>>>>>>> f7fe8a7 (Updating Branch)
intervals = (pd.DataFrame(columns=['StartTimestamp'],
                  index=pd.date_range(str(idf.iloc[0,0]) + 'T00:00:00', str(idf.iloc[0,1]) + 'T23:59:00',
                                      freq='15T'))
       .between_time('00:00','23:59')
       .index.strftime('%H:%M:%S'))
#print(intervals)

<<<<<<< HEAD
dates_df = pd.DataFrame (dates, columns = ['StartDate'])
intervals_df = pd.DataFrame (intervals, columns = ['StartTimestamp'])

#print(timestamps)
df_ids = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_DI, usecols = 'F')
consIDs = []
for x in range(1,3):
  consID = str(df_ids.loc[x])
  i = 0
  while (i < 2976):
    idSeg = consID + "-" + str(i)
    consIDs.append(idSeg)
    i+= 1

consumptionIDs_df = pd.DataFrame (consIDs, columns = ['MeterConsumptionID']) 
df_timestamps = pd.concat([consumptionIDs_df, dates_df, intervals_df], axis=1)
df_timestamps.to_excel(newSheet, sheet_name= 'DATE_INTERVAL', na_rep='', float_format=None, columns=None, header=True, index=True, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)
=======
#MeterConsumptionID:
meterConsumpID = pd.DataFrame(columns = ['Meter Consumption ID'])

# INCOMPLETE
#for x in range(2880):
#    idf.loc[meterConsumpID,'Meter Consumption ID'] = (meterConsumpID + '-' + x).zfill(len(4))

dates_df = pd.DataFrame (dates, columns = ['start'])
intervals_df = pd.DataFrame (intervals, columns = ['StartTimestamp'])
id_key = pd.DataFrame(meterConsumpID, columns = ['Meter Consumption ID'])

#print(timestamps)
df_timestamps = pd.concat([dates_df, intervals_df, id_key], axis=1)

df_timestamps.to_csv('DI.csv', index = None, header = True)

#df_timestamps.to_excel(newSheet, sheet_name= 'DATE_INTERVAL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)
#idf.to_excel(newSheet, sheet_name= 'DATE_INTERVAL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=1, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)
>>>>>>> f7fe8a7 (Updating Branch)

# --------------------------

#Modifying the sheet for POWERED_BY table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_PB, usecols = specialCols_PB)
#print(df)
df.to_excel(newSheet, sheet_name= 'POWERED_BY', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

<<<<<<< HEAD

=======
>>>>>>> f7fe8a7 (Updating Branch)
# --------------------------

#Modifying the sheet for BUILDING table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_B, usecols = specialCols_B)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'BUILDING', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for BUILDING_TYPE table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_BT, usecols = specialCols_BT)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'BUILDING_TYPE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ENERGY_SOURCE
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_ES, usecols = specialCols_ES)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'ENERGY_SOURCE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ENERGY_SOURCE_COST table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_ESC, usecols = specialCols_ESC)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'ENERGY_SOURCE_COST', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for FUEL_OIL table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_FO, usecols = specialCols_FO)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'FUEL_OIL', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for NATURAL_GAS table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_NG, usecols = specialCols_NG)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'NATURAL_GAS', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for ELECTRIC_GRID table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_EG, usecols = specialCols_EG)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'ELECTRIC_GRID', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for OTHER_SOURCE table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_OS, usecols = specialCols_OS)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'OTHER_SOURCE', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

#Modifying the sheet for MAPS_TO table
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], sheet_name = sheetName_MT, usecols = specialCols_MT)
<<<<<<< HEAD
df.drop(df.columns[0], axis=1)
=======
>>>>>>> f7fe8a7 (Updating Branch)
#print(df)
df.to_excel(newSheet, sheet_name= 'MAPS_TO', na_rep='', float_format=None, columns=None, header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None, storage_options=None)

newSheet.save()
<<<<<<< HEAD
>>>>>>>> f7fe8a7 (Updating Branch):doc/5a/Scripts/convert-CSV-collab.py
=======
>>>>>>> f7fe8a7 (Updating Branch)
